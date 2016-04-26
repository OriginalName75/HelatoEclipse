"""
    The ''Excelusing'' module
    ======================
    
    It define the functions allowing the manipulation of an Excel workbook.
    
    :Exemple:
    >> saveall()
    save every table of the database on a an Excel workbook
    
@author: IWIMBDSL
"""

from _datetime import datetime, timedelta
from datetime import date
from math import floor
import os
from time import time
import django
django.setup()
import openpyxl
from BDD.models import Modification, Personne, Groupe, TypeCour, UV, Module,\
    Note, Salle, ChampsModifie, Cour


def savmod(Modification, sheet):
    """
        Save one table of type Modification on the Excel document.
        
        :param Modification: the Modification you want to save
        :type Modification: Modification
        :param sheet: the sheet on which is the table you want to save
        :type sheet: Excel sheet
        
        :example:
        >> savpers(7,"Modification")
        save the Modification 7 on the sheet "Modification".
    """
    
    m = sheet.max_row
    i=1
    while (sheet.cell(row=i,column=1).value != None and i<m+2):
        i=i+1
    sheet.cell(row=i,column=1).value = Modification.datemodif
    sheet.cell(row=i,column=2).value = Modification.id
    sheet.cell(row=i,column=3).value = Modification.typetable
    sheet.cell(row=i,column=4).value = Modification.typemod
    sheet.cell(row=i,column=5).value = Modification.ipmod

def savchmod(ChampsModifie, sheet):
    """
        Save one table of type ChampsModifie on the Excel document.
        
        :param ChampsModifie: the ChampsModifie you want to save
        :type ChampsModifie: ChampsModifie
        :param sheet: the sheet on which is the table you want to save
        :type sheet: Excel sheet
        
        :example:
        >> savpers(name,"ChampsModifie")
        save the ChampsModifie name on the sheet "ChampsModifie".
    """
    
    m = sheet.max_row
    i=1
    while (sheet.cell(row=i,column=1).value != None and i<m+2):
        i=i+1
    sheet.cell(row=i,column=1).value = ChampsModifie.datemodif
    sheet.cell(row=i,column=2).value = ChampsModifie.idchamp
    sheet.cell(row=i,column=3).value = ChampsModifie.nomchamp
    sheet.cell(row=i,column=4).value = ChampsModifie.valchamp

def savpers(Personne, sheet):
    """
        Save one table of type Personne on the Excel document.
        
        :param Personne: the Personne you want to save
        :type Personne: Personne
        :param sheet: the sheet on which is the table you want to save
        :type sheet: Excel sheet
        
        :example:
        >> savpers(Max,"Personne")
        save the Personne Max on the sheet "Personne".
    """
    
    m = sheet.max_row
    i=1
    while (sheet.cell(row=i,column=1).value != None and i<m+2):
        i=i+1
    sheet.cell(row=i,column=1).value = Personne.uploadDate
    sheet.cell(row=i,column=2).value = Personne.id
    sheet.cell(row=i,column=3).value = Personne.user.first_name+' '+Personne.user.last_name
    sheet.cell(row=i,column=4).value = Personne.user.email
    sheet.cell(row=i,column=5).value = Personne.sexe
    sheet.cell(row=i,column=6).value = Personne.adresse
    sheet.cell(row=i,column=7).value = Personne.dateDeNaissance
    sheet.cell(row=i,column=8).value = Personne.lieuDeNaissance
    sheet.cell(row=i,column=9).value = Personne.numeroDeTel
    sheet.cell(row=i,column=10).value = Personne.type
    sheet.cell(row=i,column=11).value = Personne.promotion
    sheet.cell(row=i,column=12).value = Personne.isvisible
   
def savgrp(Groupe, sheet):
    """
        Save one table of type Groupe on the Excel document.
        
        :param Groupe: the Groupe you want to save
        :type Groupe: Groupe
        :param sheet: the sheet on which is the table you want to save
        :type sheet: Excel sheet
        
        :example:
        >> savpers(SPID,"Groupe")
        save the Groupe SPID on the sheet "Groupe".
    """
    
    m = sheet.max_row
    i=1
    while (sheet.cell(row=i,column=1).value != None and i<m+2):
        i=i+1
    pers = ''
    for p in Groupe.personnes.all():
        pers = pers+str(p)+' '
    sheet.cell(row=i,column=1).value = Groupe.uploadDate
    sheet.cell(row=i,column=2).value = Groupe.id
    sheet.cell(row=i,column=3).value = pers
    sheet.cell(row=i,column=4).value = Groupe.nom
    sheet.cell(row=i,column=5).value = Groupe.isvisible
    
def savcour(TypeCour, sheet):
    """
        Save one table of type TypeCour on the Excel document.
        
        :param TypeCour: the TypeCour you want to save
        :type TypeCour: TypeCour
        :param sheet: the sheet on which is the table you want to save
        :type sheet: Excel sheet
        
        :example:
        >> savpers(Maths,"TypeCour")
        save the TypeCour Maths on the sheet "TypeCour".
    """
    
    m = sheet.max_row
    i=1
    prof = ''
    for p in TypeCour.profs.all():
        prof = prof+str(p)+' '
    grp = ''
    for g in TypeCour.groupe.all():
        grp = grp+str(g)+' '
    while (sheet.cell(row=i,column=1).value != None and i<m+2):
        i=i+1
    sheet.cell(row=i,column=1).value = TypeCour.uploadDate
    sheet.cell(row=i,column=2).value = TypeCour.id
    sheet.cell(row=i,column=3).value = TypeCour.nom
    sheet.cell(row=i,column=4).value = prof
    sheet.cell(row=i,column=5).value = grp
    sheet.cell(row=i,column=6).value = TypeCour.isExam
    sheet.cell(row=i,column=7).value = TypeCour.isvisible
    
def savcour2(Cour, sheet):
    """
        Save one table of type Cour on the Excel document.
        
        :param Cour: the Cour you want to save
        :type Cour: Cour
        :param sheet: the sheet on which is the table you want to save
        :type sheet: Excel sheet
        
        :example:
        >> savpers(Maths,"Cour")
        save the Cour Maths on the sheet "Cour".
    """
    
    m = sheet.max_row
    i=1
    salle = ''
    for p in Cour.salles.all():
        salle = salle+str(p)+' '
    while (sheet.cell(row=i,column=1).value != None and i<m+2):
        i=i+1
    sheet.cell(row=i,column=1).value = Cour.uploadDate
    sheet.cell(row=i,column=2).value = Cour.id
    sheet.cell(row=i,column=3).value = Cour.typeCour
    sheet.cell(row=i,column=4).value = Cour.jour
    sheet.cell(row=i,column=5).value = Cour.semaineMin
    sheet.cell(row=i,column=6).value = Cour.semaineMax
    sheet.cell(row=i,column=7).value = Cour.hmin
    sheet.cell(row=i,column=8).value = Cour.hmax
    sheet.cell(row=i,column=9).value = Cour.salle
    sheet.cell(row=i,column=10).value = Cour.isvisible
    
def savuv(UV, sheet):
    """
        Save one table of type UV on the Excel document.
        
        :param UV: the UV you want to save
        :type UV: UV
        :param sheet: the sheet on which is the table you want to save
        :type sheet: Excel sheet
        
        :example:
        >> savpers(4.1,"UV")
        save the UV 4.1 on the sheet "UV".
    """
    
    m = sheet.max_row
    i=1
    while (sheet.cell(row=i,column=1).value != None and i<m+2):
        i=i+1
    sheet.cell(row=i,column=1).value = UV.uploadDate
    sheet.cell(row=i,column=2).value = UV.id
    sheet.cell(row=i,column=3).value = UV.nom

def savmodu(Module, sheet):
    """
        Save one table of type Module on the Excel document.
        
        :param Module: the Module you want to save
        :type Module: Module
        :param sheet: the sheet on which is the table you want to save
        :type sheet: Excel sheet
        
        :example:
        >> savpers(Python,"Module")
        save the Module Python on the sheet "Module".
    """
    
    m = sheet.max_row
    i=1
    while (sheet.cell(row=i,column=1).value != None and i<m+2):
        i=i+1
    sheet.cell(row=i,column=1).value = Module.uploadDate
    sheet.cell(row=i,column=2).value = Module.id
    sheet.cell(row=i,column=3).value = Module.nom
    sheet.cell(row=i,column=4).value = Module.uv.nom
    sheet.cell(row=i,column=5).value = Module.isvisible
    
def savnot(Note, sheet):
    """
        Save one table of type Note on the Excel document.
        
        :param Note: the Note you want to save
        :type Note: Note
        :param sheet: the sheet on which is the table you want to save
        :type sheet: Excel sheet
        
        :example:
        >> savpers(15,"Note")
        save the Note 15 on the sheet "Note".
    """
    
    m = sheet.max_row
    i=1
    while (sheet.cell(row=i,column=1).value != None and i<m+2):
        i=i+1
    sheet.cell(row=i,column=1).value = Note.uploadDate
    sheet.cell(row=i,column=2).value = Note.id
    sheet.cell(row=i,column=3).value = Note.note
    sheet.cell(row=i,column=4).value = Note.personne.user.first_name+' '+Note.personne.user.last_name
    sheet.cell(row=i,column=5).value = Note.module.nom
    sheet.cell(row=i,column=6).value = Note.isvisible
    
def savsal(Salle, sheet):
    """
        Save one table of type Salle on the Excel document.
        
        :param Salle: the Salle you want to save
        :type Salle: Salle
        :param sheet: the sheet on which is the table you want to save
        :type sheet: Excel sheet
        
        :example:
        >> savpers(F201,"Salle")
        save the Salle F201 on the sheet "Salle".
    """
    
    m = sheet.max_row
    i=1
    while (sheet.cell(row=i,column=1).value != None and i<m+2):
        i=i+1
    sheet.cell(row=i,column=1).value = Salle.uploadDate
    sheet.cell(row=i,column=2).value = Salle.id
    sheet.cell(row=i,column=3).value = Salle.nom
    sheet.cell(row=i,column=4).value = Salle.capacite
    sheet.cell(row=i,column=5).value = Salle.type
    sheet.cell(row=i,column=6).value = Salle.isvisible

def clearall(dat):
    """
        Clear all the data older than the date dat in the Excel document.
        
        :param dat: the date from which you begin to clear.
        :type dat: datetime
        
        :example:
        >> clearall(datetime.now())
        clear all the data stored in the Excel document that have been add before now.
    """
    
    module_dir = os.path.dirname(__file__)
    file_path = os.path.join(module_dir, 'sauvegarde.xlsx')
    file_path2 = os.path.join(module_dir, 'tempo.xlsx')
    wb = openpyxl.load_workbook(file_path)
    typ=datetime.now
    for p in wb.get_sheet_names():
        sheet = wb.get_sheet_by_name(p)
        mc = sheet.max_column
        mr = sheet.max_row
        for i in range (2,mr+1):
            if (isinstance(sheet.cell(row=i,column=1).value,float)):
                typ=convfdat(sheet.cell(row=i,column=1).value)
            else:
                typ=sheet.cell(row=i,column=1).value
            if (typ!=None and typ < dat):
                for j in range (1,mc+1):
                    sheet.cell(row=i,column=j).value = None
    wb.save(file_path2)
    wb2 = openpyxl.load_workbook(file_path2)
    wb2.save(file_path)
    
def clearmod(dat):
    """
        Clear all the Modifications older than the date dat in the Excel document.
        
        :param dat: the date from which you begin to clear.
        :type dat: datetime
        
        :example:
        >> clearall(datetime.now())
        clear all the Modifications stored in the Excel document that have been add before now.
    """
    
    module_dir = os.path.dirname(__file__)
    file_path = os.path.join(module_dir, 'sauvegarde.xlsx')
    file_path2 = os.path.join(module_dir, 'tempo.xlsx')
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.get_sheet_by_name('Modification')
    mc = sheet.max_column
    mr = sheet.max_row
    for i in range (2,mr+1):
        if (isinstance(sheet.cell(row=i,column=1).value,float)):
            typ=convfdat(sheet.cell(row=i,column=1).value)
        else:
            typ=sheet.cell(row=i,column=1).value
        if (typ!=None and typ < dat):
            for j in range (1,mc+1):
                sheet.cell(row=i,column=j).value = None
    wb.save(file_path2)
    wb2 = openpyxl.load_workbook(file_path2)
    wb2.save(file_path)

def convfdat(a):
    """
        Convert the date in the Excel document in a datetime that can be used by DJango.
        
        :param a: the date express as a float
        :type a: Float
        
        :example:
        >> convfdat(45135.45987)
        Convert the date 45135.45987 into the date 2023-07-28 11:02:13.
    """
    
    exord = datetime.toordinal(date(1900,1,1))-2
    d = datetime.fromordinal(floor(a)+exord)
    sec=round(((a-floor(a))*10**9)/11574)
    d4=d+timedelta(seconds=sec)
    return(d4)

def saveall():
    """
        Save all the table of the data base on the Excel document.
        
    """
    
    t=time()
    clearall(datetime.now())
    module_dir = os.path.dirname(__file__)
    file_path = os.path.join(module_dir, 'sauvegarde.xlsx')
    file_path2 = os.path.join(module_dir, 'tempo.xlsx')
    wb = openpyxl.load_workbook(file_path)
    svallpers(wb)
    svallgrp(wb)
    svallcour(wb)
    svallcour2(wb)
    svalluv(wb)
    svallmodu(wb)
    svallnot(wb)
    svallsal(wb)
    svallmod(wb)
    wb.save(file_path2)
    wb2 = openpyxl.load_workbook(file_path2)
    wb2.save(file_path)
    t=time()-t
    print('done in', t, 'sec')
    
def svallpers(wb):
    """
        Save all the table of type Personne on the Excel document.
        
        :param wb: the worksheet (Excel File) on which you want to save the data
        :type wb: worksheet
        
        :example:
        >> svallpers("Sauvegarde")
        save all the Personne on the Excel document "Sauvegarde".
    """
    
    sheet = wb.get_sheet_by_name('Personne')
    for p in Personne.objects.all():
        savpers(p, sheet)
         
def svallgrp(wb):
    """
        Save all the table of type Groupe on the Excel document.
        
        :param wb: the worksheet (Excel File) on which you want to save the data
        :type wb: worksheet
        
        :example:
        >> svallpers("Sauvegarde")
        save all the Groupe on the Excel document "Sauvegarde".
    """
    
    sheet = wb.get_sheet_by_name('Groupe')
    for g in Groupe.objects.all():
        savgrp(g, sheet)
        
def svallcour2(wb):
    """
        Save all the table of type Cour on the Excel document.
        
        :param wb: the worksheet (Excel File) on which you want to save the data
        :type wb: worksheet
        
        :example:
        >> svallpers("Sauvegarde")
        save all the Cour on the Excel document "Sauvegarde".
    """
    
    sheet = wb.get_sheet_by_name('Cour')
    for c in Cour.objects.all():
        savcour2(c, sheet)

def svallcour(wb):
    """
        Save all the table of type TypeCour on the Excel document.
        
        :param wb: the worksheet (Excel File) on which you want to save the data
        :type wb: worksheet
        
        :example:
        >> svallpers("Sauvegarde")
        save all the TypeCour on the Excel document "Sauvegarde".
    """
    
    sheet = wb.get_sheet_by_name('TypeCour')
    for c in TypeCour.objects.all():
        savcour(c, sheet)

def svalluv(wb):
    """
        Save all the table of type Personne on the Excel document.
        
        :param wb: the worksheet (Excel File) on which you want to save the data
        :type wb: worksheet
        
        :example:
        >> svallpers("Sauvegarde")
        save all the Personne on the Excel document "Sauvegarde".
    """
    
    sheet = wb.get_sheet_by_name('UV')
    for u in UV.objects.all():
        savuv(u, sheet)
        
def svallmodu(wb):
    """
        Save all the table of type Module on the Excel document.
        
        :param wb: the worksheet (Excel File) on which you want to save the data
        :type wb: worksheet
        
        :example:
        >> svallpers("Sauvegarde")
        save all the Module on the Excel document "Sauvegarde".
    """
    
    sheet = wb.get_sheet_by_name('Module')
    for m in Module.objects.all():
        savmodu(m, sheet)
        
def svallnot(wb):
    """
        Save all the table of type Note on the Excel document.
        
        :param wb: the worksheet (Excel File) on which you want to save the data
        :type wb: worksheet
        
        :example:
        >> svallpers("Sauvegarde")
        save all the Note on the Excel document "Sauvegarde".
    """
    
    sheet = wb.get_sheet_by_name('Note')
    for n in Note.objects.all():
        savnot(n, sheet)
        
def svallsal(wb):
    """
        Save all the table of type Salle on the Excel document.
        
        :param wb: the worksheet (Excel File) on which you want to save the data
        :type wb: worksheet
        
        :example:
        >> svallpers("Sauvegarde")
        save all the Salle on the Excel document "Sauvegarde".
    """
    
    sheet = wb.get_sheet_by_name('Salle')
    for s in Salle.objects.all():
        savsal(s, sheet)
        
def svallmod(wb):
    """
        Save all the table of type Modification on the Excel document.
        
        :param wb: the worksheet (Excel File) on which you want to save the data
        :type wb: worksheet
        
        :example:
        >> svallpers("Sauvegarde")
        save all the Modification on the Excel document "Sauvegarde".
    """
    
    sheet = wb.get_sheet_by_name('Modification')
    for m in Modification.objects.all():
        savmod(m, sheet)
    sheet = wb.get_sheet_by_name('ChampsModifie')
    for c in ChampsModifie.objects.all():
        savchmod(c, sheet)
